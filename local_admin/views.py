from django.shortcuts import render, HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm, SetPasswordForm
from django.contrib.auth.decorators import user_passes_test
from django.utils.datastructures import MultiValueDictKeyError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from datetime import datetime
from pytz import timezone
from openpyxl import Workbook
import time

from loads.forms import DateFilter, AddKeysForm
from loads.models import KeyOwner, Database, Cistern, UpDosed, Connect
from loads.views import full_recalc, refr_keys, add_key, dev,refr_loads
from .forms import EditKeyOwnerForm, AddSystemUserForm, EditDjangoUserForm, CisternForm, AddUpDosedForm

kiev = timezone('Europe/Kiev')


# Добавление резервуара
@user_passes_test(lambda u: u.is_superuser)
def cistern_add(request):
    if request.method == 'POST':
        add_cist = CisternForm(request.POST)
        try:
            if add_cist.is_valid():
                add_cist.save()
                messages.add_message(request, messages.SUCCESS, 'Резервуар успешно добавлен')
                return HttpResponseRedirect('/admin/')
        except ValueError:
            messages.add_message(request, messages.ERROR, 'Начальное значение не может превышать максимальное')
            return HttpResponseRedirect('/admin/')
    else:
        add_cist = CisternForm()
    return render(request, 'local_admin/cistern_add.html', {"add_cist": add_cist, 'cur_user': request.user.username})


@user_passes_test(lambda u: u.is_superuser)
def cistern_edit(request, cist_id):
    c = Cistern.objects.get(id=cist_id)
    if request.method == 'POST':
        edit_cist = CisternForm(request.POST, instance=c)
        try:
            if edit_cist.is_valid():
                edit_cist.save()
                if Database.objects.count() > 0:
                    full_recalc(request)
                messages.add_message(request, messages.SUCCESS, 'Значения успешно изменены')
                return HttpResponseRedirect('/admin/')
        except ValueError:
            messages.add_message(request, messages.ERROR, 'Начальное значение не может превышать максимальное')
            return HttpResponseRedirect('/admin/')
    else:
        edit_cist = CisternForm(instance=c)
    return render(request, 'local_admin/cistern_edit.html',
                  {'cur_user': request.user.username, 'edit_cist': edit_cist})


@user_passes_test(lambda u: u.is_superuser)
def keys(request):
    filter_name = ''
    filter_car = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    udb = KeyOwner.objects.filter(car__icontains=filter_car, name__icontains=filter_name)
    if request.GET.get('refr_keys'):
        refr_keys(request)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="users_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Комментарий']
        ws.append(xls_mas)
        for entry in udb:
            xls_mas = [entry.name, entry.car, entry.keys, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    if request.method == 'POST':
        add_keys_form = AddKeysForm(request.POST, request.FILES)
        try:
            key_doc = request.FILES['key_file'].read().decode().split()
        except MultiValueDictKeyError:
            messages.add_message(request, messages.ERROR, 'Ошибка файла')
            return HttpResponseRedirect('/admin/keys/')
        add_key(request, key_doc)
    else:
        add_keys_form = AddKeysForm()
    return render(request, 'local_admin/keys.html',
                  {'udb': udb, 'cur_user': request.user.username, 'add_keys': add_keys_form})


# Редактирование БД ключей
@user_passes_test(lambda u: u.is_superuser)
def edit_key(request, key_id):
    owner = KeyOwner.objects.get(id=key_id)
    if request.method == 'POST':
        edit_form = EditKeyOwnerForm(request.POST, instance=owner)
        if edit_form.is_valid():
            edit_form.save()
            messages.add_message(request, messages.SUCCESS, 'Информация успешно отредактирована')
            return HttpResponseRedirect('/admin/keys/')
    else:
        edit_form = EditKeyOwnerForm(instance=owner)
    return render(request, 'local_admin/key_edit.html',
                  {'edit_form': edit_form, 'cur_user': request.user.username})


# Пользователи системы
@user_passes_test(lambda u: u.is_superuser)
def users(request):
    last_name = ''
    if request.GET.get('last_name'):
        last_name = request.GET.get('last_name')
    udb = User.objects.filter(last_name__icontains=last_name)
    disable = request.POST.get('user_name')
    if disable:
        User.objects.get(username=disable).delete()
    return render(request, 'local_admin/users.html', {'udb': udb, 'cur_user': request.user.username})


# Регистрация пользователя
@user_passes_test(lambda u: u.is_superuser)
def add_local_user(request):
    if request.method == 'POST':
        general = UserCreationForm(request.POST)
        additional = AddSystemUserForm(request.POST)
        if general.is_valid() and additional.is_valid():
            e_mail = additional.cleaned_data['email']
            try:
                User.objects.get(email=e_mail)
                messages.add_message(request, messages.ERROR,
                                     'Пользователь с почтовым ящиком ' + e_mail + ' уже зарегистрирован')
            except User.DoesNotExist:
                user = general.save()
                user.first_name = additional.cleaned_data['first_name']
                user.last_name = additional.cleaned_data['last_name']
                user.email = e_mail
                user.save()
                messages.add_message(request, messages.SUCCESS,
                                     'Пользователь успешно зарегистрирован')
                return HttpResponseRedirect('/admin/users/')
    else:
        general = UserCreationForm()
        additional = AddSystemUserForm()
    return render(request, 'local_admin/user_add.html',
                  {'general': general, 'additional': additional, 'cur_user': request.user.username})


# Редактирование пользователя
@user_passes_test(lambda u: u.is_superuser)
def edit_local_user(request, user_id):
    u = User.objects.get(id=user_id)
    if request.method == 'POST':
        edit_form = EditDjangoUserForm(request.POST, instance=u)
        set_passwd = SetPasswordForm(user=u)
        if request.POST.get('ch_user'):
            if edit_form.is_valid():
                edit_form.save()
                messages.add_message(request, messages.SUCCESS, 'Информация о пользователе успешно отредактирована.')
        if request.POST.get('set_passwd'):
            set_passwd = SetPasswordForm(data=request.POST, user=u)
            if set_passwd.is_valid():
                set_passwd.save()
                messages.add_message(request, messages.SUCCESS, 'Пароль успешно изменен.')
    else:
        edit_form = EditDjangoUserForm(instance=u)
        set_passwd = SetPasswordForm(user=u)
    return render(request, 'local_admin/user_edit.html',
                  {'edit_form': edit_form, 'cur_user': request.user.username,
                   'e_user': u.username, 'set_passwd': set_passwd})


@user_passes_test(lambda u: u.is_superuser)
def cistern_info(request):
    nav = "downdosed"
    try:
        cist = Cistern.objects.latest('start_volume')
    except Cistern.DoesNotExist:
        messages.add_message(request, messages.ERROR,
                             'Нет зарегистрированных резервуаров. Добавьте резервуар для продолжения работы.')
        return HttpResponseRedirect('/admin/cisterns/add/')
    try:
        db = Database.objects.latest('date_time')
        vol = db.cistern_volume
    except Database.DoesNotExist:
        vol = cist.start_volume
    perc = int(vol / cist.max_volume * 100 + 5 - (vol / cist.max_volume * 100 + 5) % 10)
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    date_filter = DateFilter(request.GET)
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    if request.GET.get('nav'):
        nav = request.GET.get('nav')
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    downdosed = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, delete=False,
                                        user__name__icontains=filter_name, user__car__icontains=filter_car)
    updosed = UpDosed.objects.filter(date_time__gte=start_date, date_time__lte=end_date)
    recovery = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, delete=True,
                                       user__name__icontains=filter_name, user__car__icontains=filter_car)
    if request.method == 'POST':
        add_updosed = AddUpDosedForm(request.POST)
        if add_updosed.is_valid():
            load = Database.objects.get(id=request.POST.get('load_id'))
            if add_updosed.cleaned_data['volume'] + load.cistern_volume > load.cistern.max_volume:
                messages.add_message(request, messages.WARNING, 'Текущий объем не может быть больше максимального')
                return HttpResponseRedirect('/admin/')
            add = UpDosed(user=request.user, cistern=cist, date_time=load.date_time,
                          volume=add_updosed.cleaned_data['volume'])
            if add_updosed.cleaned_data['comment']:
                add.comment = add_updosed.cleaned_data['comment']
            add.save()
            load.add += add_updosed.cleaned_data['volume']
            load.cistern_volume += add_updosed.cleaned_data['volume']
            load.save()
            next_cist_dosings = Database.objects.filter(cistern=load.cistern,
                                                        date_time__gt=load.date_time).order_by('date_time')
            if len(next_cist_dosings):
                previous_cist_volume = load.cistern_volume
                for db in next_cist_dosings:
                    db.cistern_volume = previous_cist_volume - db.dosed + db.add
                    previous_cist_volume = db.cistern_volume
                    db.save()
            messages.add_message(request, messages.SUCCESS, 'Загрузка успешно добавлена')
    else:
        add_updosed = AddUpDosedForm()
    if request.GET.get('delete'):
        downdosed.update(delete=True)
    if request.GET.get('recover'):
        recovery.update(delete=False)
    if request.GET.get('to_xls') and nav == "downdosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'downdosed.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Объем']
        ws.append(xls_mas)
        for entry in downdosed:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys,
                       entry.dosed, entry.date_time, entry.cistern_volume]
            ws.append(xls_mas)
        wb.save(response)
        return response
    if request.GET.get('to_xls') and nav == "updosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'updosed.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Фамилия', 'Объем', 'Дата', 'Комментарий']
        ws.append(xls_mas)
        for entry in updosed:
            xls_mas = [entry.user.first_name, entry.user.last_name, entry.volume, entry.date_time, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    if request.GET.get('refr_log'):
        refr_loads(request)
    # Пагинация
    db_paginator = Paginator(downdosed, 25)
    db_page = request.GET.get('db_page')
    try:
        db = db_paginator.page(db_page)
    except PageNotAnInteger:
        db = db_paginator.page(1)
    except EmptyPage:
        db = db_paginator.page(db_paginator.num_pages)

    ud_paginator = Paginator(updosed, 25)
    ud_page = request.GET.get('ud_page')
    try:
        ud = ud_paginator.page(ud_page)
    except PageNotAnInteger:
        ud = ud_paginator.page(1)
    except EmptyPage:
        ud = ud_paginator.page(ud_paginator.num_pages)

    rec_paginator = Paginator(recovery, 25)
    rec_page = request.GET.get('rec_page')
    try:
        rec = rec_paginator.page(rec_page)
    except PageNotAnInteger:
        rec = rec_paginator.page(1)
    except EmptyPage:
        rec = rec_paginator.page(rec_paginator.num_pages)
    return render(request, 'local_admin/cistern_info.html',
                  {'cur_user': request.user.username, 'nav': nav, 'date_filter': date_filter,
                   'downdosed': db, 'updosed': ud, 'recovery': rec,
                   'add_updosed': add_updosed, 'cist': cist, 'vol': vol, 'perc': perc})



@user_passes_test(lambda u: u.is_superuser)
def hide_admin(request):
    if request.POST.get('delete_db'):
        Connect.objects.all().delete()
        UpDosed.objects.all().delete()
        Cistern.objects.all().delete()
        Database.objects.all().delete()
        KeyOwner.objects.all().delete()
        messages.add_message(request, messages.SUCCESS, 'Все данные успешно удалены.')
        return HttpResponseRedirect('/admin/hide-admin/')
    if request.POST.get('clean_log'):
        try:
            conn = Connect.objects.first()
        except Connect.DoesNotExist:
            messages.add_message(request, messages.ERROR, 'Задайте параметры подключения.')
            return HttpResponseRedirect('/admin/settings/')
        dev.port = conn.port
        dev.baudrate = conn.speed
        if not dev.isOpen():
            try:
                dev.open()
            except OSError:
                messages.add_message(request, messages.ERROR,
                                     'Невозможно установить соединение c ' + dev.port +
                                     '. Проверьте подключение устройства к компьютеру')
                return HttpResponseRedirect('/admin/')
            repeat = 0
            while repeat < 2:
                for log_on in range(len('clion\r')):
                    dev.write('clion\r'[log_on].encode())
                resp = dev.readline()
                if 'DOSA-10W' in resp.decode():
                    for log_on in range(len('clean log -f\r')):
                        dev.write('clean log -f\r'[log_on].encode())
                    resp = dev.readline()
                    if 'DOSA-10W' in resp.decode():
                        for log_on in range(len('clean log -f\r')):
                            dev.write('clean log -f\r'[log_on].encode())
                    for log_off in range(len('clioff\r')):
                        dev.write('clioff\r'[log_off].encode())
                    dev.readline()
                    dev.close()
                    repeat = 2
                else:
                    time.sleep(2)
                    repeat += 1
                    if repeat == 2:
                        dev.close()
                        messages.add_message(request, messages.ERROR,
                                             'Обмен данными c ' + dev.port +
                                             ' невозможен. Проверьте соединение устройства с RS и '
                                             'выполните подключение')
                        return HttpResponseRedirect('/admin/')
        return HttpResponseRedirect('/admin/')
    return render(request, 'local_admin/hide_admin.html')