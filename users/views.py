from django.shortcuts import render, HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from openpyxl import Workbook
from datetime import datetime
from pytz import timezone

from loads.forms import DateFilter
from loads.models import KeyOwner, Database, Cistern, UpDosed
from loads.views import refr_keys, refr_loads

kiev = timezone('Europe/Kiev')


@login_required
def cistern_info(request):
    nav = "downdosed"
    try:
        cist = Cistern.objects.latest('start_volume')
    except Cistern.DoesNotExist:
        messages.add_message(request, messages.ERROR,
                             'Нет зарегистрированных резервуаров. Обратитесь к администратору.')
        return HttpResponseRedirect('/user/keys/')
    try:
        db = Database.objects.latest('date_time')
        vol = db.cistern_volume
    except Database.DoesNotExist:
        vol = cist.start_volume
    perc = int(vol / cist.max_volume * 100 + 5 - (vol / cist.max_volume * 100 + 5) % 10)
    start_date = kiev.localize(datetime.utcfromtimestamp(0))
    end_date = kiev.localize(datetime.now())
    filter_name = ''
    filter_car = ''
    date_filter = DateFilter(request.GET)
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    if request.GET.get('nav'):
        nav = request.GET.get('nav')
    if date_filter.is_valid():
        if date_filter.cleaned_data['start_date']:
            start_date = kiev.localize(datetime.combine(date_filter.cleaned_data['start_date'], datetime.min.time()))
        if date_filter.cleaned_data['end_date']:
            end_date = kiev.localize(datetime.combine(date_filter.cleaned_data['end_date'], datetime.max.time()))
    downdosed = Database.objects.filter(date_time__gte=start_date, date_time__lte=end_date, delete=False,
                                        user__name__icontains=filter_name, user__car__icontains=filter_car)
    updosed = UpDosed.objects.filter(date_time__gte=start_date, date_time__lte=end_date)
    if request.GET.get('to_xls') and nav == "downdosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'downdosed.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Отгружено', 'Дата', 'Объем']
        ws.append(xls_mas)
        for entry in downdosed:
            xls_mas = [entry.user.name, entry.user.car, entry.user.keys,
                       entry.dosed, entry.date_time, entry.cistern_volume]
            ws.append(xls_mas)
        wb.save(response)
        return response
    if request.GET.get('to_xls') and nav == "updosed":
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = 'updosed.xls"'
        response['Content-Disposition'] = 'attachment; filename="' + filename
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Фамилия', 'Объем', 'Дата', 'Комментарий']
        ws.append(xls_mas)
        for entry in updosed:
            xls_mas = [entry.user.first_name, entry.user.last_name, entry.volume, entry.date_time, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    if request.GET.get('refr_log'):
        refr_loads(request)
    # Пагинация
    db_paginator = Paginator(downdosed, 25)
    db_page = request.GET.get('db_page')
    try:
        db = db_paginator.page(db_page)
    except PageNotAnInteger:
        db = db_paginator.page(1)
    except EmptyPage:
        db = db_paginator.page(db_paginator.num_pages)

    ud_paginator = Paginator(updosed, 25)
    ud_page = request.GET.get('ud_page')
    try:
        ud = ud_paginator.page(ud_page)
    except PageNotAnInteger:
        ud = ud_paginator.page(1)
    except EmptyPage:
        ud = ud_paginator.page(ud_paginator.num_pages)

    return render(request, 'users/cistern_user.html',
                  {'cur_user': request.user.username, 'nav': nav, 'date_filter': date_filter,
                   'downdosed': db, 'updosed': ud, 'cist': cist, 'vol': vol, 'perc': perc})


@login_required
def keys(request):
    filter_name = ''
    filter_car = ''
    if request.GET.get('filter_name'):
        filter_name = request.GET.get('filter_name')
    if request.GET.get('filter_car'):
        filter_car = request.GET.get('filter_car')
    udb = KeyOwner.objects.filter(car__icontains=filter_car, name__icontains=filter_name)
    if request.GET.get('refr_keys'):
        refr_keys(request)
    if request.GET.get('to_xls'):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="users_report.xls"'
        wb = Workbook()
        ws = wb.active
        xls_mas = ['Имя', 'Машина', 'Ключ', 'Комментарий']
        ws.append(xls_mas)
        for entry in udb:
            xls_mas = [entry.name, entry.car, entry.keys, entry.comment]
            ws.append(xls_mas)
        wb.save(response)
        return response
    return render(request, 'users/keys_user.html', {'udb': udb, 'cur_user': request.user.username})
